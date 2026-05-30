"""
Part 2: Data Classification Pipeline
=====================================
Student: Nabila Kader | ID: 23079506
Uses keyword-based ISIC classification (fast, no API needed)
"""

import sqlite3
import shutil
import sys
import os
from pathlib import Path

# QDA file extensions
QDA_EXTENSIONS = {
    'qdpx', 'nvp', 'nvivo', 'qda', 'atlproj', 'atl', 'mx24', 'mx4',
    'mx', 'qpdx', 'qde', 'f4p', 'f4', 'weft', 'transana', 'heu',
    'qda1', 'qdas'
}

# Primary data file extensions
PRIMARY_EXTENSIONS = {
    'pdf', 'doc', 'docx', 'txt', 'rtf', 'odt', 'html', 'htm',
    'mp3', 'mp4', 'wav', 'avi', 'mov', 'jpg', 'jpeg', 'png',
    'xlsx', 'xls', 'csv', 'ppt', 'pptx', 'zip'
}

def determine_project_type(project_id, conn):
    rows = conn.execute(
        "SELECT file_type FROM FILES WHERE project_id=? AND status='SUCCEEDED'",
        (project_id,)
    ).fetchall()
    
    if not rows:
        all_rows = conn.execute("SELECT file_type FROM FILES WHERE project_id=?", (project_id,)).fetchall()
        return "NOT_A_PROJECT" if not all_rows else "NOT_A_PROJECT"

    file_types = {(r[0] or "").lower().strip() for r in rows}
    if file_types & QDA_EXTENSIONS:
        return "QDA_PROJECT"
    if file_types & PRIMARY_EXTENSIONS:
        return "QD_PROJECT"
    if file_types:
        return "OTHER_PROJECT"
    return "NOT_A_PROJECT"

def classify_project(title, description, keywords):
    text = f"{title} {description or ''} {keywords or ''}".lower()
    
    rules = [
        (["hiv", "aids", "antiretroviral", "malaria", "tuberculosis", " tb ", "mortality",
          "morbidity", "epidemic", "pandemic", "covid", "vaccine", "immunization",
          "maternal health", "child health", "nutrition survey", "demographic health",
          "dhs ", "birth", "fertility", "household health", "alpha network", "incidence",
          "clinical", "disease", "hospital", "medical", "patient", "cancer", "health survey"],
         "86", "Human health activities", "72", "Scientific research and development"),

        (["education", "school", "learning", "student", "teacher", "literacy",
          "reading", "grade ", "curriculum", "classroom", "university", "higher education",
          "oer", "open educational", "early childhood", "egrs", "early grade reading",
          "training", "academic"],
         "85", "Education", "72", "Scientific research and development"),

        (["labour force", "labor force", "employment", "unemployment", "wage",
          "lfs ", "qlfs", "quarterly labour", "labour market", "labor market",
          "earnings", "occupation", "mesebetsi", "labour dynamics", "lmdsa",
          "workforce", "job seekers", "self-employed"],
         "78", "Employment activities", "84", "Public administration and defence"),

        (["agriculture", "agricultural", "farm", "farming", "crop", "livestock",
          "agra ", "food security", "smallholder", "rural household", "tobacco farmer",
          "green revolution", "irrigation", "harvest"],
         "01", "Crop and animal production", "72", "Scientific research and development"),

        (["income", "expenditure", "poverty", "financial diary", "household income",
          "consumption", "living conditions", "ies ", "lcs ", "household expenditure",
          "budget", "remittance", "microfinance"],
         "64", "Financial service activities", "88", "Social work activities without accommodation"),

        (["community survey", "census", "living standard", "general household", "ghs ",
          "nids ", "household panel", "socioeconomic", "household survey"],
         "84", "Public administration and defence", "72", "Scientific research and development"),

        (["crime", "victim", "victimisation", "safety", "police", "justice",
          "corruption", "violence", "gender-based violence", "gbv", "abuse",
          "victims of crime", "vcs "],
         "84", "Public administration and defence", "88", "Social work activities without accommodation"),

        (["energy", "electricity", "load survey", "dels ", "domestic electrical",
          "solar", "renewable", "fuel"],
         "35", "Electricity, gas, steam and air conditioning supply", None, None),

        (["water supply", "sanitation", "waste", "environment", "ecological"],
         "36", "Water collection, treatment and supply", None, None),

        (["media", "attitude", "opinion", "social attitudes", "sasas",
          "afrobarometer", "omnibus", "political culture", "election", "democracy",
          "social giving", "amps ", "all media", "product survey", "market research"],
         "73", "Advertising and market research", "84", "Public administration and defence"),

        (["tourism", "travel", "domestic tourism", "visitor", "hotel"],
         "55", "Accommodation", "93", "Sports activities and amusement and recreation activities"),

        (["manufacturing", "firm survey", "investment climate", "productivity",
          "enterprise survey", "business survey"],
         "82", "Office administrative and business support activities", "72", "Scientific research and development"),

        (["migration", "migrant", "population", "demographic", "census", "community survey"],
         "84", "Public administration and defence", "72", "Scientific research and development"),

        (["social work", "welfare", "disability", "elderly", "orphan",
          "vulnerable", "social development", "social grant"],
         "88", "Social work activities without accommodation", "84", "Public administration and defence"),
    ]

    for rule in rules:
        kw_list = rule[0]
        p_code, p_name = rule[1], rule[2]
        s_code = rule[3] if len(rule) > 3 else None
        s_name = rule[4] if len(rule) > 4 else None
        if any(kw in text for kw in kw_list):
            return (p_code, p_name, s_code, s_name)

    return ("72", "Scientific research and development", None, None)

def main():
    print("=" * 60)
    print("  Part 2: Classification Pipeline")
    print("  Student: Nabila Kader | ID: 23079506")
    print("=" * 60)

    source_db = Path("metadata.db")
    class_db = Path("23079506-sq26-classification.db")

    if class_db.exists():
        class_db.unlink()
        print(f"  Removed old {class_db}")

    print(f"\n[Step 1] Copying database to {class_db}...")
    shutil.copy2(source_db, class_db)
    print(f"  Copied.")

    conn = sqlite3.connect(class_db)

    # Add columns
    for col in ["type", "primary_class", "secondary_class"]:
        try:
            conn.execute(f"ALTER TABLE PROJECTS ADD COLUMN {col} TEXT")
            print(f"  Added '{col}' column")
        except sqlite3.OperationalError:
            pass
    conn.commit()

    # Step 2: Project type classification
    print("\n[Step 2] Classifying project types...")
    projects = conn.execute("SELECT id, title FROM PROJECTS").fetchall()
    type_counts = {"QDA_PROJECT": 0, "QD_PROJECT": 0, "OTHER_PROJECT": 0, "NOT_A_PROJECT": 0}

    for i, (pid, title) in enumerate(projects):
        ptype = determine_project_type(pid, conn)
        conn.execute("UPDATE PROJECTS SET type=? WHERE id=?", (ptype, pid))
        type_counts[ptype] += 1
        if (i + 1) % 200 == 0:
            conn.commit()
            print(f"  {i+1}/{len(projects)}...")

    conn.commit()
    print(f"\n  Project type distribution:")
    for t, c in type_counts.items():
        print(f"    {t}: {c}")

    # Step 3: ISIC classification
    print("\n[Step 3] ISIC classification (keyword-based)...")
    rows = conn.execute("""
        SELECT p.id, p.title, p.description,
               GROUP_CONCAT(k.keyword, ', ') as keywords
        FROM PROJECTS p
        LEFT JOIN KEYWORDS k ON p.id = k.project_id
        GROUP BY p.id
    """).fetchall()

    print(f"  Classifying {len(rows)} projects...")
    for i, (pid, title, description, keywords) in enumerate(rows):
        p_code, p_name, s_code, s_name = classify_project(title, description, keywords)
        primary = f"{p_code}: {p_name}"
        secondary = f"{s_code}: {s_name}" if s_code else None
        conn.execute("UPDATE PROJECTS SET primary_class=?, secondary_class=? WHERE id=?",
                    (primary, secondary, pid))
        if (i + 1) % 200 == 0:
            conn.commit()
            print(f"  {i+1}/{len(rows)}...")

    conn.commit()
    print("  Classification complete!")

    # Step 4: Export XLSX
    print("\n[Step 4] Exporting XLSX...")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Classification Results"
        headers = ["repository_id", "project_type", "project_title",
                   "primary_class", "secondary_class", "no_project_files"]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        data_rows = conn.execute("""
            SELECT p.repository_id, p.type, p.title,
                   p.primary_class, p.secondary_class,
                   COUNT(f.id) as file_count
            FROM PROJECTS p
            LEFT JOIN FILES f ON f.project_id = p.id
            GROUP BY p.id
            ORDER BY p.repository_id, p.type, p.title
        """).fetchall()

        for row in data_rows:
            ws.append(list(row))

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        xlsx_path = Path("23079506-classification.xlsx")
        wb.save(xlsx_path)
        print(f"  Saved: {xlsx_path} ({len(data_rows)} rows)")
    except ImportError:
        print("  Install openpyxl: pip install openpyxl")

    # Print statistics
    print("\n[Statistics] By repository and project type:")
    for row in conn.execute("""
        SELECT r.name, p.type, COUNT(*) as cnt
        FROM PROJECTS p JOIN REPOSITORIES r ON p.repository_id = r.id
        GROUP BY r.name, p.type ORDER BY r.name, p.type
    """):
        print(f"  {row[0]:20s} | {row[1]:20s} | {row[2]}")

    print("\n[Statistics] Top 15 primary classes:")
    for row in conn.execute("""
        SELECT primary_class, COUNT(*) as cnt FROM PROJECTS
        WHERE primary_class IS NOT NULL
        GROUP BY primary_class ORDER BY cnt DESC LIMIT 15
    """):
        print(f"  {row[1]:5d}  {row[0]}")

    conn.close()
    print(f"\n✅ Done!")
    print(f"   Database: {class_db}")
    print(f"   Spreadsheet: 23079506-classification.xlsx")

if __name__ == "__main__":
    main()
